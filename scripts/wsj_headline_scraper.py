"""
WSJ Archive Headline Scraper (v2)
==================================
Extracts headlines from WSJ daily archive pages by parsing the
__NEXT_DATA__ JSON blob embedded in the Next.js page source.

Date range: January 1, 2025 → April 4, 2026

Usage:
    pip install requests beautifulsoup4 openpyxl
    python wsj_headline_scraper_v2.py
"""

import requests
from bs4 import BeautifulSoup
import csv
import time
import random
import json
import os
from datetime import datetime, timedelta

# ─── Configuration ───────────────────────────────────────────────────────────
START_DATE = datetime(2018, 1, 1)
END_DATE = datetime(2026, 4, 4)
BASE_URL = "https://www.wsj.com/news/archive/{year}/{month:02d}/{day:02d}"
OUTPUT_CSV = "wsj_headlines.csv"
OUTPUT_XLSX = "wsj_headlines.xlsx"
CHECKPOINT_FILE = "wsj_checkpoint.json"

MIN_DELAY = 2.0
MAX_DELAY = 5.0
MAX_RETRIES = 3

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/121.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
    "Referer": "https://www.wsj.com/news/archive",
}


# ─── Helper Functions ────────────────────────────────────────────────────────

def generate_dates(start, end):
    current = start
    while current <= end:
        yield current
        current += timedelta(days=1)


def build_url(date):
    return BASE_URL.format(year=date.year, month=date.month, day=date.day)


def load_checkpoint():
    if os.path.exists(CHECKPOINT_FILE):
        with open(CHECKPOINT_FILE, "r") as f:
            data = json.load(f)
            return set(data.get("completed_dates", []))
    return set()


def save_checkpoint(completed_dates):
    with open(CHECKPOINT_FILE, "w") as f:
        json.dump({"completed_dates": sorted(completed_dates)}, f)


def extract_from_next_data(html, date):
    """
    Extract headlines from the __NEXT_DATA__ JSON blob.
    WSJ archive pages are Next.js apps that embed article data in:
      __NEXT_DATA__.props.pageProps.newsArchiveArticles
    """
    soup = BeautifulSoup(html, "html.parser")
    date_str = date.strftime("%Y-%m-%d")
    articles = []

    # Find the __NEXT_DATA__ script tag
    next_data_tag = soup.find("script", id="__NEXT_DATA__")
    if not next_data_tag or not next_data_tag.string:
        return articles

    try:
        data = json.loads(next_data_tag.string)
    except json.JSONDecodeError:
        return articles

    # Navigate to the articles array
    news_articles = (
        data.get("props", {})
            .get("pageProps", {})
            .get("newsArchiveArticles", [])
    )

    for item in news_articles:
        headline = item.get("headline", "").strip()
        url = item.get("articleUrl", "")
        timestamp = item.get("timestamp", "")

        if not headline:
            continue

        # Extract section from URL path
        # e.g. https://www.wsj.com/politics/policy/... -> politics
        section = ""
        if url:
            parts = url.replace("https://www.wsj.com/", "").split("/")
            if len(parts) >= 1 and parts[0] not in ("articles",):
                section = parts[0]

        articles.append({
            "date": date_str,
            "headline": headline,
            "url": url,
            "section": section,
            "timestamp": timestamp,
        })

    return articles


def fetch_page(url, session):
    for attempt in range(MAX_RETRIES):
        try:
            resp = session.get(url, headers=HEADERS, timeout=30)
            if resp.status_code == 200:
                return resp.text
            elif resp.status_code == 403:
                print(f"  ⚠ 403 Forbidden — may need cookies or got rate-limited")
                if attempt < MAX_RETRIES - 1:
                    time.sleep(30)
                return None
            elif resp.status_code == 429:
                wait = (attempt + 1) * 30
                print(f"  ⚠ Rate limited (429). Waiting {wait}s...")
                time.sleep(wait)
            elif resp.status_code == 404:
                print(f"  ℹ 404 — No archive page for this date")
                return None
            else:
                print(f"  ⚠ HTTP {resp.status_code} on attempt {attempt + 1}")
                time.sleep(5)
        except requests.RequestException as e:
            print(f"  ✗ Request error: {e}")
            time.sleep(5)
    return None


def write_csv(articles, filepath):
    with open(filepath, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f, fieldnames=["date", "headline", "url", "section", "timestamp"]
        )
        writer.writeheader()
        writer.writerows(articles)


def write_xlsx(articles, filepath):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "WSJ Headlines"

        header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor="1a237e")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(bottom=Side(style="thin", color="CCCCCC"))

        headers = ["Date", "Headline", "URL", "Section", "Timestamp"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        data_font = Font(name="Arial", size=10)
        link_font = Font(name="Arial", size=10, color="0563C1", underline="single")
        for i, article in enumerate(articles, 2):
            ws.cell(row=i, column=1, value=article["date"]).font = data_font
            ws.cell(row=i, column=2, value=article["headline"]).font = data_font
            ws.cell(row=i, column=3, value=article["url"]).font = link_font
            ws.cell(row=i, column=4, value=article["section"]).font = data_font
            ws.cell(row=i, column=5, value=article["timestamp"]).font = data_font
            for col in range(1, 6):
                ws.cell(row=i, column=col).border = thin_border

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 80
        ws.column_dimensions["C"].width = 70
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 24
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:E{len(articles) + 1}"

        wb.save(filepath)
        print(f"✓ Saved Excel: {filepath}")
    except ImportError:
        print("⚠ openpyxl not installed. Skipping .xlsx output.")


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    print("=" * 70)
    print("  WSJ Archive Headline Scraper v2 (__NEXT_DATA__ parser)")
    print(f"  Date range: {START_DATE.date()} → {END_DATE.date()}")
    print("=" * 70)

    all_dates = list(generate_dates(START_DATE, END_DATE))
    total_days = len(all_dates)
    print(f"  Total archive days to scrape: {total_days}")

    completed = load_checkpoint()
    if completed:
        print(f"  Resuming — {len(completed)} days already done.")

    session = requests.Session()
    all_articles = []
    new_completed = set(completed)

    # Load existing partial results
    if os.path.exists(OUTPUT_CSV) and completed:
        with open(OUTPUT_CSV, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            all_articles = list(reader)
        print(f"  Loaded {len(all_articles)} existing headlines from CSV.")

    scraped_count = 0
    for i, date in enumerate(all_dates):
        date_key = date.strftime("%Y-%m-%d")

        if date_key in completed:
            continue

        url = build_url(date)
        pct = (i + 1) / total_days * 100
        print(f"\n[{i+1}/{total_days}] ({pct:.1f}%) {date_key}  →  {url}")

        html = fetch_page(url, session)
        if html:
            headlines = extract_from_next_data(html, date)
            print(f"  ✓ Found {len(headlines)} headlines")
            all_articles.extend(headlines)
        else:
            print(f"  — No data for {date_key}")

        new_completed.add(date_key)
        scraped_count += 1

        # Checkpoint every 10 pages
        if scraped_count % 10 == 0:
            save_checkpoint(new_completed)
            write_csv(all_articles, OUTPUT_CSV)
            print(f"  💾 Checkpoint saved ({len(all_articles)} total headlines)")

        delay = random.uniform(MIN_DELAY, MAX_DELAY)
        time.sleep(delay)

    # Final save
    save_checkpoint(new_completed)
    write_csv(all_articles, OUTPUT_CSV)
    write_xlsx(all_articles, OUTPUT_XLSX)

    print("\n" + "=" * 70)
    print(f"  DONE! {len(all_articles)} headlines across {total_days} days.")
    print(f"  CSV:  {OUTPUT_CSV}")
    print(f"  XLSX: {OUTPUT_XLSX}")
    print("=" * 70)


if __name__ == "__main__":
    main()
